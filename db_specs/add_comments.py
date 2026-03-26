"""
테이블/컬럼 COMMENT 추가 스크립트
엑셀 테이블정의서에서 한글 코멘트를 읽어 PostgreSQL에 적용한다.
psycopg v3를 사용하며, DB 접속정보는 backend/.env에서 로드한다.
"""
import os
import openpyxl
import psycopg
from dotenv import load_dotenv

# backend/.env 파일에서 DB 접속정보 로드
env_path = os.path.join(os.path.dirname(__file__), "..", "backend", ".env")
load_dotenv(env_path)

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME", "tpk_db")
DB_USER = os.getenv("DB_USER", "tpk")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

DATABASE_URL = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

# 엑셀 파일 경로 — 스크립트와 같은 디렉토리의 최신 명세서 사용
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "tpk_table_spec_20260321.xlsx")

wb = openpyxl.load_workbook(EXCEL_PATH)

conn = psycopg.connect(DATABASE_URL, autocommit=True)
cur = conn.cursor()

total = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # 첫 행에서 테이블 코멘트 추출 (형식: "tb_xxx : 테이블설명")
    first_cell = str(ws.cell(row=1, column=1).value)
    table_comment = first_cell.split(":")[1].strip() if ":" in first_cell else ""

    if table_comment:
        sql = f"COMMENT ON TABLE {sheet_name} IS %s"
        cur.execute(sql, (table_comment,))
        total += 1

    # 3행부터 컬럼 데이터 (2행은 헤더)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        col_name = row[0]
        col_comment = row[1]
        if col_name and col_comment:
            sql = f"COMMENT ON COLUMN {sheet_name}.{col_name} IS %s"
            cur.execute(sql, (str(col_comment),))
            total += 1

print(f"COMMENT {total}개 적용 완료")

# 검증: 테이블 코멘트 확인
cur.execute("""
    SELECT c.relname AS table_name,
           pg_catalog.obj_description(c.oid, 'pg_class') AS table_comment
      FROM pg_catalog.pg_class c
      JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
     WHERE n.nspname = 'public'
       AND c.relkind = 'r'
     ORDER BY c.relname
""")
print("\n=== 테이블 코멘트 확인 ===")
for row in cur.fetchall():
    print(f"  {row[0]}: {row[1]}")

# 검증: 샘플 컬럼 코멘트 확인 (tb_user)
cur.execute("""
    SELECT a.attname AS column_name,
           pg_catalog.col_description(a.attrelid, a.attnum) AS column_comment
      FROM pg_catalog.pg_attribute a
      JOIN pg_catalog.pg_class c ON c.oid = a.attrelid
      JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
     WHERE n.nspname = 'public'
       AND c.relname = 'tb_user'
       AND a.attnum > 0
       AND NOT a.attisdropped
     ORDER BY a.attnum
""")
print("\n=== tb_user 컬럼 코멘트 확인 ===")
for row in cur.fetchall():
    print(f"  {row[0]}: {row[1]}")

cur.close()
conn.close()
